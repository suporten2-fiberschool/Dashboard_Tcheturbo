from __future__ import annotations

from pathlib import Path
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "progresso.xlsx"
OUTPUT_FILE = BASE_DIR / "relatorio_colaboradores_formatado.xlsx"


def normalize_nota(value: object) -> object:
    if pd.isna(value):
        return ""
    return value


def load_data(input_file: Path) -> pd.DataFrame:
    df = pd.read_excel(input_file, sheet_name="Relatório")

    # Consolida possíveis duplicidades do relatório mantendo os maiores indicadores.
    nota_rank = (
        df["Nota Prova"]
        .fillna("")
        .astype(str)
        .replace({"Pendente": "", "Curso sem prova": ""})
    )
    df = df.assign(_nota_rank=nota_rank)
    df = df.sort_values(
        by=["Nome do Aluno", "trilha", "Curso", "Progresso (%)", "Pontos Totais", "_nota_rank"],
        ascending=[True, True, True, False, False, False],
    )
    df = df.drop_duplicates(subset=["Nome do Aluno", "trilha", "Curso"], keep="first")

    return df[
        ["Nome do Aluno", "trilha", "Curso", "Progresso (%)", "Pontos Totais", "Nota Prova"]
    ].sort_values(by=["Nome do Aluno", "trilha", "Curso"], kind="stable")


def build_workbook(df: pd.DataFrame) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Formatado"

    headers = ["Nome do Aluno", "Trilha", "Curso", "Progresso (%)", "Pontuação", "Nota"]
    ws.append(headers)

    teal_fill = PatternFill("solid", fgColor="6F98A1")
    blue_fill = PatternFill("solid", fgColor="AFC3E3")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(row_number: int) -> None:
        for cell in ws[row_number]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for cell in ws[row_number][:3]:
            cell.fill = teal_fill

        for cell in ws[row_number][3:]:
            cell.fill = blue_fill

    style_header(1)

    current_row = 2
    grouped = df.groupby(["Nome do Aluno", "trilha"], sort=False)
    previous_aluno = None
    for (aluno, trilha), group in grouped:
        if previous_aluno is not None and aluno != previous_aluno:
            for col in range(1, 7):
                separator_cell = ws.cell(row=current_row, column=col, value="")
                separator_cell.fill = white_fill
                separator_cell.border = border
            current_row += 1
            for col, header in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col, value=header)
            style_header(current_row)
            current_row += 1

        cursos = group.reset_index(drop=True)
        for idx, row in cursos.iterrows():
            ws.cell(row=current_row, column=1, value=aluno if idx == 0 else "")
            ws.cell(row=current_row, column=2, value=trilha if idx == 0 else "")
            ws.cell(row=current_row, column=3, value=row["Curso"])
            ws.cell(row=current_row, column=4, value=row["Progresso (%)"])
            ws.cell(row=current_row, column=5, value=row["Pontos Totais"])
            ws.cell(row=current_row, column=6, value=normalize_nota(row["Nota Prova"]))
            current_row += 1

        previous_aluno = aluno

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    widths = {
        "A": 38,
        "B": 48,
        "C": 68,
        "D": 12,
        "E": 12,
        "F": 12,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.sheet_view.showGridLines = True

    return wb


def main() -> None:
    input_file = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else INPUT_FILE
    output_file = Path(sys.argv[2]).expanduser() if len(sys.argv) > 2 else OUTPUT_FILE

    df = load_data(input_file)
    wb = build_workbook(df)
    wb.save(output_file)
    print(f"Arquivo gerado em: {output_file}")
    print(f"Linhas exportadas: {len(df)}")


if __name__ == "__main__":
    main()
